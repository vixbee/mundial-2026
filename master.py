#!/usr/bin/env python3
"""Construye el Excel maestro.

No vuelve a simular el torneo: lee wcdata.json, que produce gen_html.py. Así el
libro y la web salen siempre de los mismos números.
"""
import os
os.chdir(os.path.dirname(os.path.abspath(__file__)) or '.')
import math, numpy as np, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
import wcbase
from wcbase import (data, EN, HOSTS, T_GOALS, DIV_F, HB_F, pts, REAL32, REAL_KO,
                    ro, r16, qf, sf, brk_meta, KOVEN, KOTIME,
                    WINNERS_G as _WIN, RUNNERS_G as _RUN, oriented_score)
from history import is_host

if not os.path.exists("wcdata.json"):
    raise SystemExit("Falta wcdata.json: corre primero `python3 gen_html.py`.")
_D = json.load(open("wcdata.json", encoding="utf-8"))
if "sim" not in _D:
    raise SystemExit("wcdata.json es de una versión anterior: vuelve a correr gen_html.py.")
SIM = _D["sim"]

mdl, ctx = wcbase.build_model()
names = ctx["names"]; idx = ctx["idx"]; NT = len(names)
team_rank = ctx["team_rank"]; team_grp = ctx["team_grp"]; groups = ctx["groups"]
RES = ctx["RES"]; XG = ctx["XG"]; SCHED = ctx["SCHED"]
N = 1.0  # las cantidades de SIM ya vienen normalizadas a probabilidad

glist = sorted(groups)
grp_teams = {g: sorted(groups[g].keys()) for g in groups}

first = np.array([SIM[nm]["pwin"] for nm in names])
second = np.array([max(SIM[nm]["padv"] - SIM[nm]["pwin"], 0.0) for nm in names])
adv = np.array([SIM[nm]["padv"] for nm in names])
third_adv = np.zeros(NT)
oct = np.array([SIM[nm]["poct"] for nm in names])
cua = np.array([SIM[nm]["pcua"] for nm in names])
sem = np.array([SIM[nm]["psem"] for nm in names])
fin = np.array([SIM[nm]["pfin"] for nm in names])
champ = np.array([SIM[nm]["pchamp"] for nm in names])


def pmf(k, l): return math.exp(-l) * l ** k / math.factorial(k)


def proj_fifa(home, hr, away, ar):
    rh = pts(hr) + (HB_F if home in HOSTS else 0)
    ra = pts(ar) + (HB_F if away in HOSTS else 0)
    S = (rh - ra) / DIV_F
    lh = max(0.18, (T_GOALS + S) / 2); la = max(0.18, (T_GOALS - S) / 2)
    ph = [pmf(i, lh) for i in range(9)]; pa = [pmf(j, la) for j in range(9)]
    best = (0, 0); bp = -1
    for i in range(9):
        for j in range(9):
            p = ph[i] * pa[j]
            if p > bp: bp = p; best = (i, j)
    return best


def model(home, hr, away, ar):
    ve = SCHED.get("|".join(sorted([home, away])), {}).get("venue", "")
    return mdl.wdl(idx[home], idx[away], is_host(home, ve), is_host(away, ve))


# Cuadro de dieciseisavos, con los clasificados reales.
def sl(t_, c): return f"{t_} ({c})"
def qc(nm): return "1.º grupo" if nm in _WIN else ("2.º grupo" if nm in _RUN else "3.º")
_meta = {mid: (dt, ve) for mid, dt, ve in brk_meta}
R32 = [(mid, _meta[mid][0], _meta[mid][1], sl(a, qc(a)), sl(b, qc(b)))
       for mid, a, b in REAL32]

ODDS = json.load(open("odds.json", encoding="utf-8"))
pmodel = {nm: SIM[nm]["pchamp"] for nm in names}
pmkt = {nm: SIM[nm]["pmkt"] for nm in names}
blend = {nm: SIM[nm]["blend"] for nm in names}
ssum = 1.0
WB = _D.get("blend_w", 0.8)
favs = [c["es"] for c in _D["champions"]]
FIT = _D.get("fit", {})
# ================= WORKBOOK =================
wb=Workbook(); FONT="Arial"
green=PatternFill("solid",fgColor="C6EFCE");yellow=PatternFill("solid",fgColor="FFEB9C");red=PatternFill("solid",fgColor="FFC7CE");blue=PatternFill("solid",fgColor="DDEBF7")
hdrf=PatternFill("solid",fgColor="1F3864");hf=Font(name=FONT,color="FFFFFF",bold=True,size=11)
thin=Side(style="thin",color="D9D9D9");bd=Border(left=thin,right=thin,top=thin,bottom=thin)
ctr=Alignment(horizontal="center",vertical="center");lft=Alignment(horizontal="left",vertical="center")
def hdr(ws,headers,row,maxcol):
    for c,h in enumerate(headers,1):
        cell=ws.cell(row=row,column=c,value=h);cell.fill=hdrf;cell.font=hf;cell.alignment=ctr;cell.border=bd

# Sheet 0: Inicio
ws0=wb.active; ws0.title="Inicio"
ws0["A1"]="Proyección Mundial 2026 — Archivo maestro"; ws0["A1"].font=Font(name=FONT,bold=True,size=18)
ws0["A2"]="Modelo basado en el ranking FIFA del 11 de junio de 2026"; ws0["A2"].font=Font(name=FONT,italic=True,size=11,color="595959")
intro=[("",""),
 ("Contenido del archivo",""),
 ("1. Partidos A-L","Proyección de los 72 partidos de grupos: probabilidad de victoria/empate/derrota, marcador más probable y ganador."),
 ("2. Avance Montecarlo","Probabilidad de cada selección de ganar su grupo y de avanzar (20.000 torneos simulados)."),
 ("3. Gráficos avance","Barras de probabilidad de avanzar por grupo."),
 ("4. Cuadro R32","Dieciseisavos emparejados con los clasificados más probables (estructura oficial FIFA)."),
 ("5. Camino a la final","Probabilidad de llegar a octavos, cuartos, semis, final y de ser campeón."),
 ("6. Favoritos","Gráfico de los 12 candidatos al título."),
 ("",""),
 ("Metodología",""),
 ("Puntos","Cada posición FIFA se convierte en puntos aproximados; la diferencia alimenta una fórmula tipo Elo (divisor 600)."),
 ("Goles","Del desbalance se derivan goles esperados y se modelan con Poisson (media 2.6 goles/partido)."),
 ("Anfitriones","Ventaja de +60 puntos solo a México, EE.UU. y Canadá en sus partidos."),
 ("Simulación","20.000 torneos completos con puntos 3/1/0, desempates oficiales y penales por fuerza relativa."),
 ("",""),
 ("Advertencias","Es una estimación probabilística reproducible, NO un pronóstico fiable. Solo usa el ranking: ignora lesiones, forma y bajas. Los cruces de octavos en adelante usan el orden estándar del cuadro."),
]
r=4
for a,b in intro:
    ca=ws0.cell(row=r,column=1,value=a); cb=ws0.cell(row=r,column=2,value=b)
    if a in ("Contenido del archivo","Metodología","Advertencias"): ca.font=Font(name=FONT,bold=True,size=12)
    else: ca.font=Font(name=FONT,bold=True,size=10)
    cb.font=Font(name=FONT,size=10); cb.alignment=Alignment(wrap_text=True,vertical="top")
    r+=1
ws0.column_dimensions["A"].width=20; ws0.column_dimensions["B"].width=95

# Sheet 1: Partidos A-L
ws=wb.create_sheet("Partidos A-L")
ws["A1"]="Resultados reales y proyección — 72 partidos de grupos (hora de Ciudad de México)"; ws["A1"].font=Font(name=FONT,bold=True,size=14); ws.merge_cells("A1:R1")
SCHED=json.load(open("sched.json",encoding="utf-8"))
H=["Grupo","Fecha","Hora (CDMX)","Estadio","Local","Rank L","Visitante","Rank V","Dif. ranking","Prob. Local","Prob. Empate","Prob. Visitante","Marcador","Pronóstico / Resultado","Ganador / Empate","Estado","Proy. original","xG (L–V)"]
hdr(ws,H,3,18); r=4
for g,date,home,hrk,away,ark in data:
    pw,pd,pl,best=model(home,hrk,away,ark)
    sc=SCHED["|".join(sorted([home,away]))]
    real=RES.get(f"{home}|{away}")
    projo="%d – %d"%proj_fifa(home,hrk,away,ark)
    xgp=XG.get(f"{home}|{away}"); xgstr=f"{xgp[0]:.2f} – {xgp[1]:.2f}" if xgp else "—"
    if real:
        gh,ga=real; marc=f"{gh} – {ga}"; estado="FINAL"
        if gh>ga: pron=f"Ganó {home}";gan=home;fav=5
        elif ga>gh: pron=f"Ganó {away}";gan=away;fav=7
        else: pron="Empate";gan="Empate";fav=None
    else:
        marc="—"; estado="Proyección"
        gap=pw-pl
        if abs(gap)<0.08: pron="Parejo";gan="Empate";fav=None
        elif gap>0: pron=f"Favorito: {home}";gan=home;fav=5
        else: pron=f"Favorito: {away}";gan=away;fav=7
    vals=[g,sc["date"],sc["time"],sc["venue"],home,hrk,away,ark,f"=ABS(F{r}-H{r})",round(pw,3),round(pd,3),round(pl,3),marc,pron,gan,estado,projo,xgstr]
    for c,v in enumerate(vals,1):
        cell=ws.cell(row=r,column=c,value=v); cell.border=bd; cell.font=Font(name=FONT,size=11)
        cell.alignment=lft if c in (4,5,7,14) else ctr
    for c in (10,11,12): ws.cell(row=r,column=c).number_format="0%"
    ws.cell(row=r,column=17).font=Font(name=FONT,size=11,italic=True,color="7F7F7F")
    if estado=="FINAL": ws.cell(row=r,column=16).font=Font(name=FONT,size=11,bold=True,color="1F3864")
    if fav is None:
        for c in (13,14,15): ws.cell(row=r,column=c).fill=yellow; ws.cell(row=r,column=c).font=Font(name=FONT,size=11,bold=True,color="9C6500")
    else:
        ws.cell(row=r,column=fav).fill=green; ws.cell(row=r,column=fav).font=Font(name=FONT,size=11,bold=True,color="006100")
        for c in (14,15): ws.cell(row=r,column=c).fill=green; ws.cell(row=r,column=c).font=Font(name=FONT,size=11,bold=True,color="006100")
    r+=1
for col,wd in {"A":7,"B":9,"C":10,"D":30,"E":20,"F":8,"G":20,"H":8,"I":12,"J":11,"K":12,"L":13,"M":12,"N":22,"O":20,"P":12,"Q":13,"R":13}.items(): ws.column_dimensions[col].width=wd
ws.freeze_panes="A4"; ws.auto_filter.ref=f"A3:R{r-1}"

# Sheet 2: Avance Montecarlo
wsA=wb.create_sheet("Avance Montecarlo")
wsA["A1"]="Probabilidad de avanzar — Montecarlo (20.000 torneos)"; wsA["A1"].font=Font(name=FONT,bold=True,size=14); wsA.merge_cells("A1:G1")
hdr(wsA,["Grupo","Selección","Rank","Prob. ganar grupo","Prob. 1º o 2º","Prob. avanzar","Prob. avanzar"],3,7); r=4; grp_rows={}
for g in glist:
    rowsg=sorted([(nm,team_rank[nm],first[idx[nm]]/N,(first[idx[nm]]+second[idx[nm]])/N,adv[idx[nm]]/N) for nm in grp_teams[g]],key=lambda x:x[4],reverse=True)
    grp_rows[g]=(r,r+len(rowsg)-1)
    for nm,rk,pw,pt,pa in rowsg:
        for c,v in enumerate([g,nm,rk,round(pw,3),round(pt,3),round(pa,3),round(pa,3)],1):
            cell=wsA.cell(row=r,column=c,value=v); cell.border=bd; cell.font=Font(name=FONT,size=11); cell.alignment=lft if c==2 else ctr
        for c in (4,5,6,7): wsA.cell(row=r,column=c).number_format="0%"
        if pa>=0.55: f,fc=green,"006100"
        elif pa>=0.25: f,fc=yellow,"9C6500"
        else: f,fc=red,"9C0006"
        for c in (2,7): wsA.cell(row=r,column=c).fill=f; wsA.cell(row=r,column=c).font=Font(name=FONT,size=11,bold=True,color=fc)
        r+=1
for col,wd in {"A":7,"B":20,"C":7,"D":17,"E":13,"F":13,"G":13}.items(): wsA.column_dimensions[col].width=wd
wsA.freeze_panes="A4"; wsA.auto_filter.ref=f"A3:G{r-1}"

# Sheet 3: Gráficos avance
wsG=wb.create_sheet("Gráficos avance")
wsG["A1"]="Probabilidad de avanzar por grupo"; wsG["A1"].font=Font(name=FONT,bold=True,size=14)
for i,g in enumerate(glist):
    r0,r1=grp_rows[g]; ch=BarChart(); ch.type="col"; ch.title=f"Grupo {g}"; ch.legend=None
    ch.y_axis.numFmt='0%'; ch.y_axis.scaling.min=0; ch.y_axis.scaling.max=1; ch.height=6.2; ch.width=10
    ch.add_data(Reference(wsA,min_col=7,min_row=r0,max_row=r1),titles_from_data=False)
    ch.set_categories(Reference(wsA,min_col=2,min_row=r0,max_row=r1))
    wsG.add_chart(ch,f"{'B' if i%2==0 else 'L'}{3+(i//2)*13}")

# Sheet 4: Cuadro R32
wsR=wb.create_sheet("Cuadro R32")
wsR["A1"]="Dieciseisavos (Round of 32) — clasificados más probables"; wsR["A1"].font=Font(name=FONT,bold=True,size=14); wsR.merge_cells("A1:E1")
hdr(wsR,["Partido","Fecha","Equipo 1 (proyectado)","Equipo 2 (proyectado)","Sede"],3,5); r=4
for mid,fe,se,e1,e2 in R32:
    for c,v in enumerate([mid,fe,e1,e2,se],1):
        cell=wsR.cell(row=r,column=c,value=v); cell.border=bd; cell.font=Font(name=FONT,size=11); cell.alignment=ctr if c in(1,2) else lft
        if c in (3,4): cell.fill=blue
    r+=1
for col,wd in {"A":9,"B":9,"C":30,"D":30,"E":26}.items(): wsR.column_dimensions[col].width=wd
wsR.freeze_panes="A4"

# Sheet 5: Camino a la final
wsK=wb.create_sheet("Camino a la final")
wsK["A1"]="Probabilidades en eliminatorias — Montecarlo"; wsK["A1"].font=Font(name=FONT,bold=True,size=14); wsK.merge_cells("A1:I1")
hdr(wsK,["Selección","Grupo","Rank","Clasifica","Octavos","Cuartos","Semifinal","Final","Campeón"],3,9)
allrows=sorted([(names[i],team_grp[names[i]],team_rank[names[i]],adv[i]/N,oct[i]/N,cua[i]/N,sem[i]/N,fin[i]/N,champ[i]/N) for i in range(NT)],key=lambda x:x[8],reverse=True)
r=4
for nm,g,rk,pcl,po,pc,ps,pf,pch in allrows:
    for c,v in enumerate([nm,g,rk,round(pcl,3),round(po,3),round(pc,3),round(ps,3),round(pf,3),round(pch,3)],1):
        cell=wsK.cell(row=r,column=c,value=v); cell.border=bd; cell.font=Font(name=FONT,size=11); cell.alignment=lft if c==1 else ctr
    for c in range(4,10): wsK.cell(row=r,column=c).number_format="0%"
    if pch>=0.15: f,fc=green,"006100"
    elif pch>=0.05: f,fc=yellow,"9C6500"
    else: f,fc=None,"000000"
    if f: wsK.cell(row=r,column=1).fill=f; wsK.cell(row=r,column=9).fill=f
    wsK.cell(row=r,column=9).font=Font(name=FONT,size=11,bold=True,color=fc)
    r+=1
for col,wd in {"A":20,"B":7,"C":7,"D":11,"E":10,"F":10,"G":11,"H":9,"I":10}.items(): wsK.column_dimensions[col].width=wd
wsK.freeze_panes="A4"; wsK.auto_filter.ref=f"A3:I{r-1}"

# Sheet 6: Favoritos (modelo con xG + mercado + mezcla)
ODDS=json.load(open("odds.json",encoding="utf-8"))
imp={nm:100.0/(o+100.0) for nm,o in ODDS.items()}; ssum=sum(imp.values())
pmkt={nm:imp[nm]/ssum for nm in imp}; WB=0.8
pmodel={names[i]:champ[i]/N for i in range(NT)}
blend={nm:WB*pmkt.get(nm,0.0)+(1-WB)*pmodel[nm] for nm in names}
bs=sum(blend.values()); blend={nm:blend[nm]/bs for nm in blend}
favs=sorted(names,key=lambda nm:blend[nm],reverse=True)[:12]
wsF=wb.create_sheet("Favoritos")
wsF["A1"]="Probabilidad de ser campeón — modelo (con xG) + mercado (cuotas)"; wsF["A1"].font=Font(name=FONT,bold=True,size=14); wsF.merge_cells("A1:D1")
wsF["A2"]="Mezcla 70% mercado / 30% modelo. Modelo = simulación condicionada a resultados reales y xG. Mercado = cuotas de campeón al 16 jun, sin margen de la casa."
wsF["A2"].font=Font(name=FONT,italic=True,size=9,color="595959"); wsF.merge_cells("A2:D2")
hdr(wsF,["Selección","Modelo","Mercado","Mezcla"],3,4)
for k,nm in enumerate(favs):
    wsF.cell(row=4+k,column=1,value=nm).alignment=lft; wsF.cell(row=4+k,column=1).border=bd
    for c,v in [(2,pmodel[nm]),(3,pmkt.get(nm,0.0)),(4,blend[nm])]:
        cell=wsF.cell(row=4+k,column=c,value=round(v,3)); cell.number_format="0%"; cell.alignment=ctr; cell.border=bd
    bl=blend[nm]
    if bl>=0.12: f,fc=green,"006100"
    elif bl>=0.05: f,fc=yellow,"9C6500"
    else: f,fc=None,None
    if f: wsF.cell(row=4+k,column=4).fill=f; wsF.cell(row=4+k,column=4).font=Font(name=FONT,size=11,bold=True,color=fc)
for col,wd in {"A":18,"B":10,"C":10,"D":10}.items(): wsF.column_dimensions[col].width=wd
ch=BarChart(); ch.type="bar"; ch.title="Probabilidad de campeón (mezcla)"; ch.legend=None; ch.x_axis.numFmt='0%'; ch.height=9; ch.width=16
ch.add_data(Reference(wsF,min_col=4,min_row=4,max_row=15),titles_from_data=False); ch.set_categories(Reference(wsF,min_col=1,min_row=4,max_row=15))
wsF.add_chart(ch,"F3")

# Sheet 7: Cuadro completo (octavos -> final)
KO=json.load(open("wcdata.json",encoding="utf-8"))["knockout"]
wsK2=wb.create_sheet("Cuadro octavos-final")
wsK2["A1"]="Proyección del cuadro — de octavos a la final (avanza el favorito del modelo)"; wsK2["A1"].font=Font(name=FONT,bold=True,size=14); wsK2.merge_cells("A1:D1")
wsK2["A2"]=("Ruta más probable partido a partido a partir del cuadro de dieciseisavos: en cada llave avanza el equipo con mayor probabilidad "
 "según el modelo (condicionado a los resultados reales al 16 jun). El equipo que avanza va en verde. Es un escenario único; su probabilidad conjunta es baja.")
wsK2["A2"].font=Font(name=FONT,italic=True,size=9,color="595959"); wsK2.merge_cells("A2:D2"); wsK2["A2"].alignment=Alignment(wrap_text=True,vertical="top"); wsK2.row_dimensions[2].height=42
def ko_block(ws,title,rows_,r0):
    c=ws.cell(row=r0,column=1,value=title); c.font=Font(name=FONT,bold=True,size=12,color="1F3864"); r0+=1
    for ci,h in enumerate(["Partido","Equipo 1","Equipo 2","Sede"],1):
        cell=ws.cell(row=r0,column=ci,value=h); cell.fill=hdrf; cell.font=hf; cell.alignment=ctr; cell.border=bd
    r0+=1
    for m in rows_:
        ws.cell(row=r0,column=1,value=m.get("mid","—")).alignment=ctr
        ws.cell(row=r0,column=2,value=m["a"]["es"]).alignment=lft
        ws.cell(row=r0,column=3,value=m["b"]["es"]).alignment=lft
        fecha_hora=m.get("date","")+(" "+m["time"]+"h" if m.get("time") else "")
        sede=(fecha_hora+" · "+m.get("venue","")).strip(" ·")
        ws.cell(row=r0,column=4,value=sede).alignment=lft
        winc=2 if m["w"]=="a" else 3
        for ci in range(1,5):
            cell=ws.cell(row=r0,column=ci); cell.border=bd; cell.font=Font(name=FONT,size=11)
        ws.cell(row=r0,column=winc).fill=green; ws.cell(row=r0,column=winc).font=Font(name=FONT,size=11,bold=True,color="006100")
        r0+=1
    return r0+1
rr=4
rr=ko_block(wsK2,"Octavos de final",KO["octavos"],rr)
rr=ko_block(wsK2,"Cuartos de final",KO["cuartos"],rr)
rr=ko_block(wsK2,"Semifinales",KO["semis"],rr)
rr=ko_block(wsK2,"Final",[KO["final"]],rr)
rr=ko_block(wsK2,"Tercer lugar",[KO["third"]],rr)
champ_cell=wsK2.cell(row=rr,column=1,value=f"CAMPEÓN PROYECTADO: {KO['champion']['es']}  ({round(KO['champion_p']*100)}% en la simulación)")
champ_cell.font=Font(name=FONT,bold=True,size=13,color="006100"); wsK2.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=4)
for col,wd in {"A":12,"B":22,"C":22,"D":34}.items(): wsK2.column_dimensions[col].width=wd

wb.save("Proyeccion-Mundial-2026-MAESTRO.xlsx")
print("Hojas:", wb.sheetnames)
